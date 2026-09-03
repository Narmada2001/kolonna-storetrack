import io
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy.orm import Session
from sqlalchemy import extract

from .. import schemas
from ..auth import require_admin
from ..database import get_db
from ..models import Item, ItemRequest, RequestStatus, Supplier, Transaction

router = APIRouter(prefix="/reports", tags=["reports"], dependencies=[Depends(require_admin)])


@router.get("/dashboard", response_model=schemas.DashboardStats)
def dashboard_stats(db: Session = Depends(get_db)):
    items = db.query(Item).all()
    low_stock = sum(1 for i in items if i.quantity_in_stock <= i.reorder_level)
    pending = db.query(ItemRequest).filter(ItemRequest.status == RequestStatus.pending).count()
    suppliers = db.query(Supplier).count()
    now = datetime.utcnow()
    txns_this_month = (
        db.query(Transaction)
        .filter(
            extract("year", Transaction.transaction_date) == now.year,
            extract("month", Transaction.transaction_date) == now.month,
        )
        .count()
    )
    return schemas.DashboardStats(
        total_items=len(items),
        low_stock_items=low_stock,
        pending_requests=pending,
        total_suppliers=suppliers,
        transactions_this_month=txns_this_month,
    )


@router.get("/dashboard/transactions", response_model=list[schemas.TransactionTimeseriesPoint])
def transactions_timeseries(days: int = 30, db: Session = Depends(get_db)):
    """Daily received vs. issued quantity for the last `days` days (usage-pattern chart)."""
    days = max(1, min(days, 365))
    today = datetime.utcnow().date()
    start = today - timedelta(days=days - 1)
    start_of_range = datetime(start.year, start.month, start.day)

    buckets = {
        (start + timedelta(days=i)).isoformat(): {"received": 0, "issued": 0} for i in range(days)
    }

    rows = db.query(Transaction).filter(Transaction.transaction_date >= start_of_range).all()
    for t in rows:
        key = t.transaction_date.date().isoformat()
        if key in buckets:
            buckets[key][t.type.value] += t.quantity

    return [
        schemas.TransactionTimeseriesPoint(date=date, received=v["received"], issued=v["issued"])
        for date, v in buckets.items()
    ]


REPORT_BUILDERS = {}


def _inventory_rows(db: Session, start_date: str = None, end_date: str = None, filter_val: str = None):
    header = ["ID", "Name", "Category", "In Stock", "Reorder Level", "Unit Price", "Low Stock?"]
    rows = [header]
    query = db.query(Item).order_by(Item.name)
    if filter_val == "low_stock":
        query = query.filter(Item.quantity_in_stock <= Item.reorder_level)
    for item in query.all():
        rows.append(
            [
                item.id,
                item.name,
                item.category or "-",
                item.quantity_in_stock,
                item.reorder_level,
                float(item.unit_price),
                "YES" if item.quantity_in_stock <= item.reorder_level else "",
            ]
        )
    return rows


def _requests_rows(db: Session, start_date: str = None, end_date: str = None, filter_val: str = None):
    header = ["ID", "Employee", "Item", "Qty", "Status", "Requested On"]
    rows = [header]
    query = db.query(ItemRequest).order_by(ItemRequest.request_date.desc())
    if start_date:
        query = query.filter(ItemRequest.request_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(ItemRequest.request_date < datetime.fromisoformat(end_date) + timedelta(days=1))
    if filter_val and filter_val != "all":
        query = query.filter(ItemRequest.status == filter_val)
        
    for r in query.all():
        rows.append(
            [
                r.id,
                r.employee.full_name if r.employee else "-",
                r.item.name if r.item else "-",
                r.quantity,
                r.status.value,
                r.request_date.strftime("%Y-%m-%d %H:%M"),
            ]
        )
    return rows


def _transactions_rows(db: Session, start_date: str = None, end_date: str = None, filter_val: str = None):
    header = ["ID", "Item", "Type", "Qty", "Supplier", "Date", "Ref/Notes", "Cost"]
    rows = [header]
    query = db.query(Transaction).order_by(Transaction.transaction_date.desc())
    if start_date:
        query = query.filter(Transaction.transaction_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Transaction.transaction_date < datetime.fromisoformat(end_date) + timedelta(days=1))
    if filter_val and filter_val != "all":
        query = query.filter(Transaction.type == filter_val)
        
    for t in query.all():
        ref = t.reference_no or ""
        notes = getattr(t, "notes", "") or ""
        ref_notes = f"{ref} {notes}".strip()
        
        rows.append(
            [
                t.id,
                (t.item.name[:20] + "...") if t.item and len(t.item.name) > 20 else (t.item.name if t.item else "-"),
                t.type.value,
                t.quantity,
                (t.supplier.name[:15] + "...") if t.supplier and len(t.supplier.name) > 15 else (t.supplier.name if t.supplier else "-"),
                t.transaction_date.strftime("%Y-%m-%d"),
                ref_notes[:25] + "..." if len(ref_notes) > 25 else (ref_notes or "-"),
                f"${float(getattr(t, 'total_cost', 0.0)):.2f}"
            ]
        )
    return rows


REPORT_BUILDERS = {
    "inventory": ("Inventory Report", _inventory_rows),
    "requests": ("Item Requests Report", _requests_rows),
    "transactions": ("Transactions Report", _transactions_rows),
}


@router.get("/{report_name}/pdf")
def export_pdf(
    report_name: str, 
    start_date: str = None, 
    end_date: str = None, 
    filter_val: str = None,
    db: Session = Depends(get_db)
):
    if report_name not in REPORT_BUILDERS:
        raise HTTPException(status_code=404, detail="Unknown report")
    title, builder = REPORT_BUILDERS[report_name]
    rows = builder(db, start_date, end_date, filter_val)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    
    total_records = len(rows) - 1
    summary = f"Total Records: {total_records}"
    if start_date and end_date:
        summary += f" | Period: {start_date} to {end_date}"
    if filter_val and filter_val != "all":
        summary += f" | Filter: {filter_val.capitalize()}"

    elements = [
        Paragraph("Kolonna StoreTrack", styles["Title"]),
        Paragraph(title, styles["Heading2"]),
        Paragraph(datetime.utcnow().strftime("Generated on %Y-%m-%d %H:%M UTC"), styles["Normal"]),
        Paragraph(summary, styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]
    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={report_name}_report.pdf"},
    )


@router.get("/{report_name}/excel")
def export_excel(
    report_name: str, 
    start_date: str = None, 
    end_date: str = None, 
    filter_val: str = None,
    db: Session = Depends(get_db)
):
    if report_name not in REPORT_BUILDERS:
        raise HTTPException(status_code=404, detail="Unknown report")
    title, builder = REPORT_BUILDERS[report_name]
    rows = builder(db, start_date, end_date, filter_val)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    
    for row in rows:
        ws.append(row)
        
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try: 
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
        
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_name}_report.xlsx"},
    )
